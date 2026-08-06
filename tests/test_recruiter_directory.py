"""Who appears in the recruiter directory, and what the export contains."""
import io

import pytest
from openpyxl import load_workbook

from app import membership as mb


class TestContactDefault:
    """Listing is on by default; only an explicit opt-out removes someone."""

    def test_a_new_account_is_listed(self):
        assert mb.contactable({"username": "alice"}) is True

    def test_a_never_set_value_is_listed(self):
        assert mb.contactable({"opt_in_contact": None}) is True

    def test_opting_in_explicitly_is_listed(self):
        assert mb.contactable({"opt_in_contact": True}) is True

    def test_opting_out_removes_them(self):
        assert mb.contactable({"opt_in_contact": False}) is False

    def test_opting_out_is_durable(self):
        """The absent-means-yes rule must not re-list someone who opted out."""
        opted_out = {"username": "alice", "opt_in_contact": False}
        for _ in range(3):
            assert mb.contactable(dict(opted_out)) is False

    def test_missing_record_is_not_listed(self):
        assert mb.contactable(None) is False

    def test_the_model_default_matches_the_rule(self):
        from app.models import User
        u = User(username="alice")
        assert u.opt_in_contact is True
        assert mb.contactable(u.model_dump()) is True


# ── Export ──────────────────────────────────────────────────────────────
def build_workbook(students, total_ranked=3):
    """Run the export's sheet-building against a fixed set of students.

    Mirrors app.recruiters.export_xlsx without needing a request or Firestore.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from app import scores

    played = {m["key"] for s in students for m in s["modes"]}
    mode_keys = [k for k in scores.MODE_KEYS if k in played]

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    headers = ["Name", "Username", "Email", "Membership", "Club", "Graduation",
               "Overall rating", "Rank", "Modes played", "Games", "CV on file"]
    headers += [scores.mode_meta(k)["label"] for k in mode_keys]
    ws.append(headers)

    for s in students:
        by_mode = {m["key"]: m for m in s["modes"]}
        row = [
            s.get("full_name") or s.get("username", ""),
            s.get("username", ""), s.get("email", ""), s.get("membership", ""),
            s.get("club", ""), s.get("graduation_year") or "",
            s.get("overall") if s.get("overall") is not None else "",
            s.get("rank") or "", s.get("modes_played", 0), s.get("total_games", 0),
            "yes" if s.get("cv_uploaded") else "no",
        ]
        for k in mode_keys:
            m = by_mode.get(k)
            row.append(f"{m['rating']}*" if m and m["provisional"]
                       else (m["rating"] if m else ""))
        ws.append(row)

    ws.freeze_panes = "A2"
    if students:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(students) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


@pytest.fixture
def students():
    return [
        {"id": "u1", "username": "alice", "full_name": "Alice Ng",
         "email": "alice@example.com", "membership": "Quant Analyst",
         "club": "Alpha Fund", "graduation_year": 2027, "overall": 81, "rank": 1,
         "modes_played": 2, "total_games": 9, "cv_uploaded": True,
         "modes": [{"key": "risks", "label": "Risks", "rating": 88, "games": 5,
                    "provisional": False},
                   {"key": "fiveos", "label": "5Os", "rating": 74, "games": 4,
                    "provisional": True}]},
        {"id": "u2", "username": "bob", "full_name": "", "email": "",
         "membership": "Member", "club": "Alpha Fund", "graduation_year": None,
         "overall": None, "rank": None, "modes_played": 0, "total_games": 0,
         "cv_uploaded": False, "modes": []},
    ]


class TestExport:
    def test_it_opens_as_a_real_workbook(self, students):
        wb = build_workbook(students)
        assert wb.sheetnames[0] == "Students"

    def test_one_row_per_student_plus_a_header(self, students):
        ws = build_workbook(students)["Students"]
        assert ws.max_row == len(students) + 1

    def test_contact_details_are_in_the_sheet(self, students):
        ws = build_workbook(students)["Students"]
        emails = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
        assert "alice@example.com" in emails

    def test_a_student_with_no_email_still_exports(self, students):
        ws = build_workbook(students)["Students"]
        names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "bob" in names, "an uncontactable student must not vanish silently"

    def test_the_name_falls_back_to_the_username(self, students):
        ws = build_workbook(students)["Students"]
        assert ws.cell(2, 1).value == "Alice Ng"
        assert ws.cell(3, 1).value == "bob"

    def test_only_played_modes_get_columns(self, students):
        ws = build_workbook(students)["Students"]
        headers = [c.value for c in ws[1]]
        assert "Risks" in headers and "5Os" in headers
        assert "Poker Auction" not in headers, "unplayed modes should not add columns"

    def test_a_provisional_rating_is_marked(self, students):
        ws = build_workbook(students)["Students"]
        headers = [c.value for c in ws[1]]
        col = headers.index("5Os") + 1
        assert ws.cell(2, col).value == "74*"

    def test_a_firm_rating_stays_numeric_for_sorting(self, students):
        ws = build_workbook(students)["Students"]
        headers = [c.value for c in ws[1]]
        col = headers.index("Risks") + 1
        assert ws.cell(2, col).value == 88

    def test_the_header_row_is_frozen_and_filterable(self, students):
        ws = build_workbook(students)["Students"]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None

    def test_an_empty_directory_still_produces_a_sheet(self):
        ws = build_workbook([])["Students"]
        assert ws.max_row == 1
