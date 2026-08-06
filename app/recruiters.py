"""
Recruiter directory.
====================

Students who opt in to being contacted appear here, with their performance and
an email address; recruiters get in touch from their own inbox. Nothing is
sent on anyone's behalf.

Two gates, both required:

* the viewer must hold the **recruiter** role (granted by an admin, on request
  — a shared code would leak this to anyone it reached), and
* the student must have **opted in**, which is off by default and reversible
  at any time from their profile.

Blacklisted accounts and staff never appear.
"""
from __future__ import annotations

import asyncio
import datetime as dt
import logging
from pathlib import Path
from typing import Any, Dict, List

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.templating import Jinja2Templates
from firebase_admin import auth as fb_auth

from app import db as db_module
from app import membership as mb
from app import scores
from app.auth import current_user
from app.models import User

log = logging.getLogger("uvicorn.error")

router = APIRouter(prefix="/recruiters", tags=["recruiters"])
BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


async def _contact_email(doc_id: str, data: Dict[str, Any]) -> str:
    """The student's address, healing the record if Firestore has not got one.

    `users.email` only started being written when this directory was built, and
    older documents fill in lazily on the owner's next sign-in. A student who
    opted in but has not signed in since would appear here with no way to reach
    them — which defeats the point. Firebase Auth has held the address all
    along, so read it from there and write it back, once.
    """
    email = (data or {}).get("email")
    if email:
        return email

    uid = str((data or {}).get("firebase_uid") or doc_id)
    try:
        email = await asyncio.to_thread(lambda: fb_auth.get_user(uid).email)
    except Exception as exc:
        log.info("recruiters: no Firebase address for %s: %s", uid, exc)
        return ""

    if not email:
        return ""
    try:
        await db_module.db.collection("users").document(doc_id).update({"email": email})
    except Exception as exc:
        log.warning("recruiters: could not cache email for %s: %s", doc_id, exc)
    return email


async def _require_recruiter(user: User) -> Dict[str, Any]:
    doc = await db_module.db.collection("users").document(str(user.id)).get()
    data = doc.to_dict() or {}
    if not mb.is_recruiter({**data, "is_admin": user.is_admin}):
        raise HTTPException(
            403,
            "The recruiter directory is open to approved recruiters. "
            "You can request the role from your profile.",
        )
    return data


@router.get("", include_in_schema=False)
async def directory_page(request: Request):
    return templates.TemplateResponse("recruiters.html", {
        "request": request,
        "app_name": "AlphaBook",
    })


async def _rows_for_recruiter() -> Dict[str, Any]:
    """Every contactable student, strongest first, with performance and contact."""
    board = await scores.leaderboard()
    ratings = {p["user_id"]: p for p in board["players"]}
    total = len(board["players"])

    docs = await db_module.db.collection("users").get()
    rows: List[Dict[str, Any]] = []
    for d in docs:
        data = d.to_dict() or {}
        if data.get("is_admin") or data.get("is_blacklisted"):
            continue
        if not mb.contactable(data):
            continue

        me = ratings.get(d.id)
        modes = []
        if me:
            for key, r in (me.get("ratings") or {}).items():
                meta = scores.mode_meta(key)
                modes.append({
                    "key": key,
                    "label": meta["label"],
                    "rating": r["rating"],
                    "games": r["games"],
                    "provisional": r["provisional"],
                })
            modes.sort(key=lambda m: m["rating"], reverse=True)

        rows.append({
            **mb.public_profile(d.id, data),
            "email": await _contact_email(d.id, data),
            "cv_uploaded": bool(data.get("cv_blob_path")),
            "in_cv_book": mb.cv_book_included(data),
            "overall": me["overall"] if me else None,
            "rank": me["rank"] if me else None,
            "modes_played": me["modes_played"] if me else 0,
            "total_games": me["total_games"] if me else 0,
            "modes": modes,
        })

    # Strongest first, then everyone still unranked.
    rows.sort(key=lambda r: (r["overall"] is None, -(r["overall"] or 0), r["username"].lower()))
    return {
        "students": rows,
        "total_ranked": total,
        "memberships": mb.MEMBERSHIPS,
    }


@router.get("/directory")
async def directory(user: User = Depends(current_user)):
    """Opted-in students, with performance and contact details."""
    await _require_recruiter(user)
    return await _rows_for_recruiter()


@router.get("/export.xlsx", include_in_schema=False)
async def export_xlsx(user: User = Depends(current_user)):
    """The directory as a spreadsheet, for working through offline."""
    from fastapi.responses import StreamingResponse
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    await _require_recruiter(user)
    data = await _rows_for_recruiter()
    students = data["students"]

    # Which modes to give columns to: every one anybody has actually played,
    # in the platform's own order, so the sheet is stable between exports.
    played = {m["key"] for s in students for m in s["modes"]}
    mode_keys = [k for k in scores.MODE_KEYS if k in played]

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["Name", "Username", "Email", "Membership", "Club", "Graduation",
               "Overall rating", "Rank", "Modes played", "Games", "CV on file"]
    headers += [scores.mode_meta(k)["label"] for k in mode_keys]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1B75BC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")

    for s in students:
        by_mode = {m["key"]: m for m in s["modes"]}
        row = [
            s.get("full_name") or s.get("username", ""),
            s.get("username", ""),
            s.get("email", ""),
            s.get("membership", ""),
            s.get("club", ""),
            s.get("graduation_year") or "",
            s.get("overall") if s.get("overall") is not None else "",
            s.get("rank") or "",
            s.get("modes_played", 0),
            s.get("total_games", 0),
            "yes" if s.get("cv_uploaded") else "no",
        ]
        for k in mode_keys:
            m = by_mode.get(k)
            # A provisional rating is marked, not hidden — the number is real
            # but thin, and a recruiter should see which is which.
            row.append(f"{m['rating']}*" if m and m["provisional"]
                       else (m["rating"] if m else ""))
        ws.append(row)

    widths = [24, 18, 30, 20, 16, 12, 14, 8, 13, 8, 11] + [16] * len(mode_keys)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if students:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(students) + 1}"

    notes = wb.create_sheet("About")
    for line in [
        ["AlphaBook — recruiter directory export"],
        [],
        ["Ratings are percentiles against every other player on the platform,"],
        ["so they compare across cohorts. 0-100, higher is better."],
        ["A * marks a provisional rating: too few games to be firm yet."],
        [],
        ["Students appear here unless they opted out in their settings."],
        ["Contact them from your own inbox; nothing is sent by AlphaBook."],
        [],
        [f"Students in this export: {len(students)}"],
        [f"Ranked players on the platform: {data['total_ranked']}"],
        [f"Generated: {dt.datetime.now(dt.timezone.utc):%Y-%m-%d %H:%M UTC}"],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="alphabook-students-{stamp}.xlsx"'},
    )


@router.get("/students/{user_id}/cv", include_in_schema=False)
async def student_cv(user_id: str, user: User = Depends(current_user)):
    """Stream an opted-in student's CV to an approved recruiter."""
    from fastapi.responses import StreamingResponse
    import asyncio
    import io

    await _require_recruiter(user)

    if not db_module.bucket:
        raise HTTPException(500, "Storage not configured")

    doc = await db_module.db.collection("users").document(user_id).get()
    data = doc.to_dict() if doc.exists else {}
    if not data or not mb.contactable(data):
        raise HTTPException(404, "No such student, or they have not opted in")

    blob_name = data.get("cv_blob_path")
    if not blob_name:
        raise HTTPException(404, "That student has not uploaded a CV")

    pdf = await asyncio.get_event_loop().run_in_executor(
        None, db_module.bucket.blob(blob_name).download_as_bytes
    )
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": 'inline; filename="cv.pdf"'},
    )
